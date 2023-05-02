from datetime import date

def sort_by_date(filtered_report):
    report = {}
    for index, item in enumerate(filtered_report):
        report.setdefault(str(item.date_create), []).append(item)
    return report


def sort_by_meal(report):
    report_ = {}
    for key, value in report.items():
        report_[key] = {}
        for index, item in enumerate(value):
            report_[key].setdefault(str(item.meal), []).append(item)
    return report_


def change_str_in_date(date_str):
    date_list = date_str.split('-')
    return date(int(date_list[0]), int(date_list[1]), int(date_list[2]))

def create_external_report(filtered_report):
    """ Отчет для Hadassah. """
    report = sort_by_date(filtered_report)

    report_ = sort_by_meal(report)

    money_total = 0
    count_total = 0
    for date_key, one_day_report in report_.items():
        report[date_key] = {}
        count_all = 0 # все рационы кроме "Нулевой диеты"
        count_just_wather = 0
        count_bd_2 = 0
        price_all = 750
        price_just_wather = 150
        price_count_bd_2 = 150

        if change_str_in_date(date_key) >= date(2023, 5, 1):
            price_bouillon = 0
        else:
            price_bouillon = 90
        for meal_key in ['breakfast', 'lunch', 'afternoon', 'dinner']:
            meal_report = one_day_report.get(meal_key, [])
            # создать функцию, которая добаляет уникальные диеты в report
            report[date_key][meal_key] = {}
            users = list(set([item_report.user_id for item_report in meal_report]))
            for user in users:
                report_user_set = [report for report in meal_report if report.user_id==user]
                if len([report_item for report_item in report_user_set if report_item.product_id=='426']) > 0 and \
                        report_user_set[0].meal == 'dinner' and report_user_set[0].type_of_diet == 'БД день 2' or \
                        len([report_item for report_item in report_user_set if report_item.product_id == '426']) == 0:
                    for report_user in report_user_set:
                        report[date_key][meal_key].setdefault(str(report_user.type_of_diet), []).append(report_user)
                else:
                    for report_user in report_user_set:
                        report[date_key][meal_key].setdefault(f'{report_user.type_of_diet} + бульон', []).append(report_user)
            for diet_key, on_diet_report in report[date_key][meal_key].items():
                count_items = len(set([user.user_id for user in (report[date_key][meal_key][diet_key])]))
                count_bouillon = \
                    len(set([user.user_id for user in (report[date_key][meal_key][diet_key])
                                if user.product_id == '426' and\
                                not (meal_key == 'dinner' and diet_key == 'БД день 2')]))
                if diet_key == 'Нулевая диета':
                    price = price_just_wather
                    count_just_wather += count_items
                elif diet_key == 'Нулевая диета + бульон':
                    price = price_just_wather
                    count_just_wather += count_items
                elif diet_key == 'БД день 2' and meal_key == 'afternoon':
                    price = price_count_bd_2
                    count_bd_2 += count_items
                else:
                    price = price_all
                count_all += count_items
                report[date_key][meal_key][diet_key] =\
                    {'count': count_items,
                     'count_bouillon': count_bouillon,
                     'money': (count_items * price) + (count_bouillon * price_bouillon)}

        money = (count_all - count_just_wather - count_bd_2) * price_all + \
                (count_just_wather * price_just_wather) + \
                (count_bd_2 * price_count_bd_2) + \
                (count_bouillon * price_bouillon)
        report[date_key]['Всего'] = {'count': count_all,
                                     'money': money}
        money_total += money
        count_total += count_all
    report['Итого'] = {'Всего за период':
                           {'count': count_total,
                            'money': money_total}}
    return report


def get_report(report, report_detailing, date_start, date_finish):
    """ Создаёт excel фаил с отчетом по блюдам """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    font = Font(name='Arial',
                size=10,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')
    font_white = Font(name='Arial',
                size=10,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='ffffff')
    font_10_bold = Font(name='Arial',
                size=10,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')
    ws['A1'].font = font
    ws['A2'].font = font
    ws['A3'].font = font
    ws['A5'].font = font
    ws['B5'].font = font
    ws['C5'].font = font
    ws['D5'].font = font
    ws['E5'].font = font
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    ws.merge_cells('A3:E3')
    ws.merge_cells('A4:E4')
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.row_dimensions[5].height = 30
    ws['A1'].value = 'Отчет по лечебному питанию'
    ws['A1'].font = Font(name='Arial',
                         size=14,
                         bold=False,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')
    ws['A2'].value = 'Круглосуточный стационар, Hadassah Medical Moscow'
    ws['A2'].font = Font(name='Arial',
                         size=11,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')
    ws['A3'].value = f'{date_start.day}.{date_start.month}.{date_start.year} - {date_finish.day}.{date_finish.month}.{date_finish.year}'
    ws['A3'].font = Font(name='Arial',
                         size=11,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')

    ws['A5'].value = 'Учетный день'
    ws['B5'].value = 'Прием пищи'
    ws['C5'].value = 'Рацион'
    ws['D5'].value = 'Количество'
    ws['E5'].value = 'Сумма, руб.'
    ws['A5'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B5'].alignment = Alignment(horizontal="left", vertical="center")
    ws['C5'].alignment = Alignment(horizontal="left", vertical="center")
    ws['D5'].alignment = Alignment(horizontal="left", vertical="center")
    ws['E5'].alignment = Alignment(horizontal="left", vertical="center")

    dotted = Side(border_style="dotted", color="383636")
    thick = Side(border_style="thin", color="383636")
    row = 5
    for key1, item1 in report.items():
        row += 1
        if key1 != 'Итого':
            _ = ws.cell(column=1, row=row, value=key1).font = font
        for key2, item2 in item1.items():
            if key2 in ['Всего','Всего за период']:
                _ = ws.cell(column=2, row=row, value=key2).font = font_10_bold
            if key2 == 'breakfast':
                _ = ws.cell(column=2, row=row, value='Завтрак').font = font
            if key2 == 'lunch':
                _ = ws.cell(column=2, row=row, value='Обед').font = font
            if key2 == 'afternoon':
                _ = ws.cell(column=2, row=row, value='Полдник').font = font
            if key2 == 'dinner':
                _ = ws.cell(column=2, row=row, value='Ужин').font = font
            for key3, item3 in item2.items():
                if key2 in ['Всего','Всего за период']:
                    _ = ws.cell(column=3, row=row, value='').font = font_10_bold
                    _ = ws.cell(column=4, row=row, value=str(item2['count'])).font = font_10_bold
                    _ = ws.cell(column=5, row=row, value=f'{item2["money"]}.00').font = font_10_bold
                else:
                    _ = ws.cell(column=3, row=row, value=key3).font = font
                    _ = ws.cell(column=4, row=row, value=str(item3['count'])).font = font
                    _ = ws.cell(column=5, row=row, value=f'{item3["money"]}.00').font = font
                    row += 1
            ws[row-1][0].border = Border(bottom=dotted)
            ws[row-1][1].border = Border(bottom=dotted)
            ws[row-1][2].border = Border(bottom=dotted)
            ws[row-1][3].border = Border(bottom=dotted)
            ws[row-1][4].border = Border(bottom=dotted)
        ws[row][0].border = Border(bottom=thick)
        ws[row][1].border = Border(bottom=thick)
        ws[row][2].border = Border(bottom=thick)
        ws[row][3].border = Border(bottom=thick)
        ws[row][4].border = Border(bottom=thick)

    for row in range(1, row+300):
        for col in range(1, 50):
            ws.cell(column=col, row=row).fill = PatternFill('solid', fgColor="ffffff")

    ws['A5'].fill = PatternFill('solid', fgColor="203864")
    ws['B5'].fill = PatternFill('solid', fgColor="203864")
    ws['C5'].fill = PatternFill('solid', fgColor="203864")
    ws['D5'].fill = PatternFill('solid', fgColor="203864")
    ws['E5'].fill = PatternFill('solid', fgColor="203864")
    ws['A5'].font = font_white
    ws['B5'].font = font_white
    ws['C5'].font = font_white
    ws['D5'].font = font_white
    ws['E5'].font = font_white

    ws1 = wb.create_sheet("Детальный")

    font = Font(name='Arial',
                size=10,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')
    ws1['A1'].font = font
    ws1['A2'].font = font
    ws1['A3'].font = font
    ws1['A5'].font = font
    ws1['B5'].font = font
    ws1['C5'].font = font
    ws1['D5'].font = font
    ws1['E5'].font = font
    ws1.merge_cells('A1:E1')
    ws1.merge_cells('A2:E2')
    ws1.merge_cells('A3:E3')
    ws1.merge_cells('A4:E4')
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 15
    ws1.row_dimensions[5].height = 30
    ws1['A1'].value = 'Детализация к отчету по лечебному питанию'
    ws1['A1'].font = Font(name='Arial',
                         size=14,
                         bold=False,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')
    ws1['A2'].value = 'Круглосуточный стационар, Hadassah Medical Moscow'
    ws1['A2'].font = Font(name='Arial',
                         size=11,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')
    ws1['A3'].value = f'{date_start.day}.{date_start.month}.{date_start.year} - {date_finish.day}.{date_finish.month}.{date_finish.year}'
    ws1['A3'].font = Font(name='Arial',
                         size=11,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')

    ws1['A5'].value = 'Учетный день'
    ws1['B5'].value = 'Прием пищи'
    ws1['C5'].value = 'Рацион'
    ws1['D5'].value = 'ФИО'
    ws1['E5'].value = '№ палаты'
    ws1['A5'].alignment = Alignment(horizontal="left", vertical="center")
    ws1['B5'].alignment = Alignment(horizontal="left", vertical="center")
    ws1['C5'].alignment = Alignment(horizontal="left", vertical="center")
    ws1['D5'].alignment = Alignment(horizontal="left", vertical="center")
    ws1['E5'].alignment = Alignment(horizontal="left", vertical="center")
    row = 6
    for key1, item1 in report_detailing.items():
        for key2, item2 in item1.items():
            for key3, item3 in item2.items():
                for key4, value4 in item3.items():
                    _ = ws1.cell(column=1, row=row, value=key1).font = font
                    _ = ws1.cell(column=1, row=row, value=key1).font = font
                    if key2 == 'breakfast':
                        _ = ws1.cell(column=2, row=row, value='Завтрак').font = font
                    if key2 == 'lunch':
                        _ = ws1.cell(column=2, row=row, value='Обед').font = font
                    if key2 == 'afternoon':
                        _ = ws1.cell(column=2, row=row, value='Полдник').font = font
                    if key2 == 'dinner':
                        _ = ws1.cell(column=2, row=row, value='Ужин').font = font
                    _ = ws1.cell(column=3, row=row, value=key3).font = font
                    _ = ws1.cell(column=4, row=row, value=key4).font = font
                    _ = ws1.cell(column=5, row=row, value=f'{value4 if value4 != "Не выбрано" else "——"}').font = font
                    row += 1


    for row in range(1, row+300):
        for col in range(1, 50):
            ws1.cell(column=col, row=row).fill = PatternFill('solid', fgColor="ffffff")


    ws1['A5'].fill = PatternFill('solid', fgColor="203864")
    ws1['B5'].fill = PatternFill('solid', fgColor="203864")
    ws1['C5'].fill = PatternFill('solid', fgColor="203864")
    ws1['D5'].fill = PatternFill('solid', fgColor="203864")
    ws1['E5'].fill = PatternFill('solid', fgColor="203864")
    ws1['A5'].font = font_white
    ws1['B5'].font = font_white
    ws1['C5'].font = font_white
    ws1['D5'].font = font_white
    ws1['E5'].font = font_white
    wb.save("static/report.xlsx")
    return



def create_external_report_detailing(filtered_report):
    report = {}
    report_ = {}
    for index, item in enumerate(filtered_report):
        report.setdefault(str(item.date_create), []).append(item)

    for key, value in report.items():
        report_[key] = {}
        for index, item in enumerate(value):
            report_[key].setdefault(str(item.meal), []).append(item)

    for key1, value1 in report_.items():
        report[key1] = {}

        for key2, value2 in value1.items():
            report[key1][key2] = {}
            for index, item in enumerate(value2):
                report[key1][key2].setdefault(str(item.type_of_diet), []).append(item)
            for key3, value3 in report[key1][key2].items():
                test = {}
                for item in report[key1][key2][key3]:
                    test[item.user_id.full_name] = item.user_id.room_number
                report[key1][key2][key3] = test
    return report
