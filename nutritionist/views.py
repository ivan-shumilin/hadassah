import json, os, requests, random, math, calendar, datetime, re, operator, openpyxl
from dateutil.parser import parse
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.models import User
from django.forms import modelformset_factory
from django.core.paginator import Paginator
from django.template import RequestContext
from django.urls import reverse
from django.db import transaction
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required, user_passes_test
from django.forms import CheckboxInput, Textarea
from django.core.mail import send_mail
from django.db.models.functions import Lower
from .models import Base, Product, Timetable, CustomUser, Barcodes, ProductLp, MenuByDay, BotChatId, СhangesUsersToday,\
    UsersToday, UsersReadyOrder, MenuByDayReadyOrder, Report
from .forms import UserRegistrationForm, UserloginForm, TimetableForm, UserPasswordResetForm
from .serializers import ProductSerializer
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.renderers import JSONRenderer
from rest_framework.views import APIView
from datetime import datetime, date
from django.core import management
from django.core.management.commands import dumpdata
from django.contrib.auth.models import Group
from doctor.functions.functions import sorting_dishes, parsing, get_day_of_the_week, translate_diet, add_default_menu, \
    creates_dict_with_menu_patients, add_menu_three_days_ahead, creating_meal_menu_lp, creating_meal_menu_cafe, \
    creates_dict_with_menu_patients_on_day, delete_choices, create_user, edit_user, check_have_menu, counting_diets, \
    create_list_users_on_floor, what_meal, translate_meal, check_value_two, what_type_order
from doctor.functions.bot import check_change
from doctor.functions.for_print_forms import create_user_today, check_time, update_UsersToday, update_СhangesUsersToday, \
    applies_changes
from doctor.functions.bot import formatting_full_name
import random, calendar, datetime, logging, json
from datetime import datetime, date, timedelta
from django.utils import dateformat


def group_nutritionists_check(user):
    return user.groups.filter(name='nutritionists').exists()


def backup(request):
    answer = []
    management.call_command('backup_db_ydisk', stdout=answer)
    return render(request, 'backup.html', {'answer': answer})


@transaction.atomic
def load_menu(dict_tests, dict_tk):
    menu_items = dict_tests['menu']['items']
    to_create = []
    for menu_item in menu_items:
        try:
            if len(menu_item['product']['description']) < 3:
                menu_item['product']['description'] = 'Отсутствует'
        except TypeError:
            menu_item['product']['description'] = 'Отсутствует'
        key = dict_tk.get(menu_item['product']['name'])
        if key == None:
            key = "Отсутствует"

        if (len(Product.objects.filter(iditem=menu_item['product']['id'])) == 0) and (
                len([item for item in to_create if item.iditem == menu_item['product']['id']]) == 0):
            to_create.append(Product(
                iditem=menu_item['product']['id'],
                name=menu_item['product']['name'],
                price=menu_item['product']['price'],
                carbohydrate=menu_item['product']['carbohydrate'],
                fat=menu_item['product']['fat'],
                fiber=menu_item['product']['fiber'],
                energy=menu_item['product']['energy'],
                image=menu_item['product']['image'],
                vegan=menu_item['product']['vegan'],
                allergens=menu_item['product']['allergens'],
                lactose_free=menu_item['product']['lactose_free'],
                sugarless=menu_item['product']['sugarless'],
                gluten_free=menu_item['product']['gluten_free'],
                description=menu_item['product']['description'],
                category=menu_item['category']['name'],
                cooking_method=key
            ))
    Product.objects.bulk_create(to_create)


@transaction.atomic
def load_timetable(dict_tests):
    # Product.objects.all().delete()
    # dict_tests = {'menu': {'id': 682, 'date': '24.06.2022', 'status': 'completed', 'completed_at': '22.06.2022 17:52:31', 'created_at': '17.06.2022 10:26:17', 'combo_price': 350, 'location': {'id': 4, 'name': 'hadassah', 'subdomain': 'hadassah'}, 'items': [{'id': 13276, 'combo': False, 'product': {'id': 251, 'name': 'Говядина по-азиатски 75/75 гр.', 'price': 239, 'carbohydrate': '0.57539', 'fat': '11.42332', 'fiber': '21.60205', 'energy': '191.51808', 'image': 'a8faed5c-fd3e-40cf-bb93-6d86240a268e.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав: говядина, масло подсолнечное, соус соевый, крахмал, соус терияки, перец чили, масло кунжутное, кунжут, корень имбиря, перец болгарский, лук репчатый.'}, 'category': {'id': 6, 'name': 'Вторые блюда'}}, {'id': 13277, 'combo': False, 'product': {'id': 337, 'name': 'Куриная ватрушка с грибным жульеном 120 гр.', 'price': 159, 'carbohydrate': '23.04720', 'fat': '17.42760', 'fiber': '20.17560', 'energy': '329.73120', 'image': '66366171-300e-469f-8692-de7da76785f4.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав : Состав : фарш из курицы ( филе курицы, хлеб, соль, перекц , лук репчатый), масло подсолнечное, сыр Гауда, жульен грибной ( грибы шампиньоны, лук репчатый, масло подсолнечное, молоко 3,2%, сливки 33%, перец черный молотый, соль).'}, 'category': {'id': 6, 'name': 'Вторые блюда'}}, {'id': 13278, 'combo': False, 'product': {'id': 396, 'name': 'Перец фаршированный овощами, грибами и кус-кусом 200/50', 'price': 239, 'carbohydrate': '25.70250', 'fat': '8.97750', 'fiber': '4.95250', 'energy': '203.40750', 'image': '7bba3000-acb3-4b5d-aecd-aa0b7be477e5.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав: перец болгарский, грибы шампиньоны, кускус, морковь, перец болгарский, томатный соус ( томатная паста, лук репчатый, морковь, перец черный молотый, соль, прованские травы),'}, 'category': {'id': 6, 'name': 'Вторые блюда'}}, {'id': 13279, 'combo': True, 'product': {'id': 1817, 'name': 'Суп-крем овощной 250 гр.', 'price': 99, 'carbohydrate': '21.50694', 'fat': '5.40708', 'fiber': '2.86319', 'energy': '146.14958', 'image': 'fb998852-3579-4a9a-bba7-abe5cccb8f32.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав: капуста брокколи, капуста цветная , лук репчатый, картофель, морковь, зхелень, сельдерей, перец, соль, масло оливковое, сливки 33 % .'}, 'category': {'id': 17, 'name': 'Первые блюда'}}, {'id': 13280, 'combo': False, 'product': {'id': 1255, 'name': 'Тайский суп с лапшой и говядиной 250 гр.', 'price': 169, 'carbohydrate': '17.91500', 'fat': '14.48250', 'fiber': '15.07500', 'energy': '262.29000', 'image': '76113578-d361-41f9-9bad-31a37e26c856.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав: шпинат, говядина, чеснок, имбирь, лук репчатый, морковь, масло подсолнечное, грибы шампиньоны, лапша гречневая, соевый соус, соль, перец.'}, 'category': {'id': 17, 'name': 'Первые блюда'}}, {'id': 13281, 'combo': False, 'product': {'id': 434, 'name': 'Стейк из бедра индейки с томатной сальсой 100/50 гр.', 'price': 239, 'carbohydrate': '3.64950', 'fat': '16.14150', 'fiber': '21.92400', 'energy': '247.57050', 'image': 'df5fb311-1b7f-4373-bc4d-2ca869c8b3e7.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав : филе индейки, соевый соус, тимьян, мед, масло подсолнечное, томатная сальса( помидор, лук репчатый, чеснок, кинза, соль, перец, масло оливковое).'}, 'category': {'id': 6, 'name': 'Вторые блюда'}}, {'id': 13282, 'combo': True, 'product': {'id': 446, 'name': 'Тефтели куриные с зеленым горошком в сметанном соусе 100/50гр.', 'price': 139, 'carbohydrate': '16.39800', 'fat': '12.20250', 'fiber': '14.38200', 'energy': '232.92450', 'image': '1b066181-48e5-4739-b847-6819afa21f54.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав: фарш куриный (курица, соль, перец, лук, хлеб),горек зеленый, сметана, мука пшеничная, соль,перец черный молотый.'}, 'category': {'id': 6, 'name': 'Вторые блюда'}}, {'id': 13283, 'combo': False, 'product': {'id': 514, 'name': 'Брокколи на пару с болгарским перцем 150 гр.', 'price': 119, 'carbohydrate': '7.41150', 'fat': '5.26650', 'fiber': '3.73650', 'energy': '91.98150', 'image': '0f3a5af0-8454-4441-a5f6-c53655b1b2b3.jpg', 'vegan': 1, 'allergens': 0, 'lactose_free': None, 'sugarless': None, 'gluten_free': None, 'description': 'Состав:Капуста брокколи с/м, перец болгарский, масло подсолнечное'}, 'category': {'id': 9, 'name': 'Гарниры'}}, {'id': 13284, 'combo': True, 'product': {'id': 617, 'name': 'Спагетти отварные 150 гр.', 'price': 69, 'carbohydrate': '35.74950', 'fat': '7.15650', 'fiber': '4.68000', 'energy': '226.12350', 'image': '66149542-2a5f-426e-9e58-bf327a581b76.jpg', 'vegan': 1, 'allergens': 0, 'lactose_free': None, 'sugarless': None, 'gluten_free': None, 'description': 'Состав Спагетти, масло подс.,соль.'}, 'category': {'id': 9, 'name': 'Гарниры'}}, {'id': 13285, 'combo': False, 'product': {'id': 359, 'name': 'Лазанья грибная 200 гр.', 'price': 199, 'carbohydrate': '56.29400', 'fat': '14.49800', 'fiber': '19.03600', 'energy': '431.79400', 'image': 'f82d609c-7824-4492-b297-a0616726f9cb.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав : грибы шампиньоны, лук репчатый, масло подсолнечное, паста для лазаньи, перец черный молотый, соль, сыр гауда, соус бешамель ( масло сливочное, молоко 3,2%, мускатный орех, мука пшеничная, соль, перец черный молотый.)'}, 'category': {'id': 6, 'name': 'Вторые блюда'}}, {'id': 13286, 'combo': False, 'product': {'id': 1218, 'name': 'Суп рисовый с куриными фрикадельками 250 гр.', 'price': 89, 'carbohydrate': '10.14000', 'fat': '4.60750', 'fiber': '6.47750', 'energy': '107.94000', 'image': 'dd6bbb7a-d17c-4725-b8b8-2da02e2012db.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав Бульон куриный, фрикадельки из куриного фарша,рис, морковь, лук репчатый,соль , перец'}, 'category': {'id': 17, 'name': 'Первые блюда'}}, {'id': 13287, 'combo': False, 'product': {'id': 508, 'name': ' Картофельное пюре со шпинатом150 гр.', 'price': 89, 'carbohydrate': '26.91450', 'fat': '7.95750', 'fiber': '4.05450', 'energy': '195.50250', 'image': 'ee560f38-5b7b-41f8-8275-a4a2b3e7f8b1.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав: Картофель,молоко,масло сливочное,шпинат с/м.,соль'}, 'category': {'id': 9, 'name': 'Гарниры'}}, {'id': 13288, 'combo': False, 'product': {'id': 1642, 'name': 'Рис с укропом150 гр.', 'price': 69, 'carbohydrate': '47.23619', 'fat': '4.42347', 'fiber': '3.91797', 'energy': '241.06339', 'image': 'fb9a613c-3095-4a6b-a194-5a695b628b90.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Cостав:Рис,шпинат,масло раст.,специи'}, 'category': {'id': 9, 'name': 'Гарниры'}}, {'id': 13289, 'combo': False, 'product': {'id': 927, 'name': 'Салат "Столичный" с курицей 100 гр.', 'price': 89, 'carbohydrate': '5.98515', 'fat': '9.44750', 'fiber': '6.03912', 'energy': '133.12458', 'image': '298406ab-7fd7-4b54-8c47-ac94ea3ac828.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав: картофель, морковь, филе куриное, огурцы маринованные, огурцы свежие, горошек консерв, яйцо куриное отварное, майонез, зелень, соль.'}, 'category': {'id': 15, 'name': 'Салаты'}}, {'id': 13290, 'combo': False, 'product': {'id': 947, 'name': 'Салат из крабовых палочек 100 гр.', 'price': 89, 'carbohydrate': '5.40000', 'fat': '9.16500', 'fiber': '2.91000', 'energy': '115.72500', 'image': '606430d1-cf3b-4d6d-ac82-050ed6bf6bd0.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав: Огурцы свежие, капуста пекинская, крабовые палочки,масйонез, кукуруза консервированная.'}, 'category': {'id': 15, 'name': 'Салаты'}}, {'id': 13291, 'combo': False, 'product': {'id': 992, 'name': 'Салат из томатов с редисом 100 гр.', 'price': 109, 'carbohydrate': '3.58000', 'fat': '5.16000', 'fiber': '0.80500', 'energy': '63.98000', 'image': '1c25edbd-0443-4865-a7e0-ba0c9417e44f.jpg', 'vegan': 0, 'allergens': 0, 'lactose_free': 0, 'sugarless': 0, 'gluten_free': 0, 'description': 'Состав:Помидоры свежие, редис, укроп, масло раст., соль.'}, 'category': {'id': 15, 'name': 'Салаты'}}, {'id': 13292, 'combo': True, 'product': {'id': 976, 'name': 'Салат из свежей капусты с зеленым горошком 100 гр.', 'price': 79, 'carbohydrate': '4.82248', 'fat': '7.27273', 'fiber': '2.26123', 'energy': '93.78938', 'image': '5f3ea239-a7b2-47e8-9de5-3ca456a6e175.jpg', 'vegan': 1, 'allergens': 0, 'lactose_free': None, 'sugarless': None, 'gluten_free': None, 'description': 'Состав:Капуста белокачанная, горошек консерв., зелень, масло раст., соль.'}, 'category': {'id': 15, 'name': 'Салаты'}}, {'id': 13293, 'combo': False, 'product': {'id': 958, 'name': 'Салат из огурцов с укропом 100 гр.', 'price': 89, 'carbohydrate': '2.87000', 'fat': '2.12600', 'fiber': '0.86200', 'energy': '34.06200', 'image': 'cc924895-5a6e-49d3-b344-13173340541a.jpg', 'vegan': 1, 'allergens': None, 'lactose_free': None, 'sugarless': None, 'gluten_free': None, 'description': 'Cостав:Огурцы свежие, укроп,масло раст.,соль'}, 'category': {'id': 15, 'name': 'Салаты'}}]}}

    menu_items = dict_tests['menu']['items']
    to_create = []
    for menu_item in menu_items:
        menu_item_date = datetime.strptime(dict_tests['menu']['date'], "%d.%m.%Y").strftime("%Y-%m-%d")
        if len(Timetable.objects.filter(datetime=menu_item_date).filter(
                item=Product.objects.get(iditem=menu_item['product']['id']))) == 0:
            to_create.append(Timetable(
                datetime=menu_item_date,
                item=Product.objects.get(iditem=menu_item['product']['id'])
            ))
    Timetable.objects.bulk_create(to_create)


def redirect(request):
    return HttpResponseRedirect(reverse('index'))


@user_passes_test(group_nutritionists_check, login_url='login')
@login_required(login_url='login')
def index(request):
    error = ''
    ProductFormSet = modelformset_factory(Product,
                                          fields=(
                                              'iditem', 'name', 'description', 'ovd', 'ovd_sugarless', 'shd', 'bd',
                                              'vbd', 'nbd', 'nkd',
                                              'vkd', 'not_suitable', 'carbohydrate', 'fat', 'fiber', 'energy', 'category',
                                              'cooking_method', 'comment'),
                                          widgets={'ovd': CheckboxInput(
                                              attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'ovd_sugarless': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'shd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'bd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'not_suitable': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'name': Textarea(attrs={'style': "display: none;"}),
                                              'description': Textarea(attrs={'style': "display: none;"}),
                                              'carbohydrate': Textarea(attrs={'style': "display: none;"}),
                                              'iditem': Textarea(attrs={'style': "display: none;"}),
                                              'fat': Textarea(attrs={'style': "display: none;"}),
                                              'fiber': Textarea(attrs={'style': "display: none;"}),
                                              'energy': Textarea(attrs={'style': "display: none;"}),
                                              'category': Textarea(attrs={'style': "display: none;"}),
                                              'cooking_method': Textarea(attrs={'style': "display: none;"}),
                                              'comment': Textarea(attrs={'class': "form-control", 'rows': "3"}),
                                          },
                                          extra=0, )
    if request.method == 'GET':
        date_default = str(date.today())
    else:
        date_default = str(request.POST['datetime'])
    count_prosucts = len(Product.objects.all())
    count_prosucts_labeled = len(Product.objects.filter(
        Q(ovd='True') | Q(ovd_sugarless='True') | Q(shd='True') | Q(bd='True') | Q(vbd='True') | Q(nbd='True') | Q(
            nkd='True') | Q(vkd='True') | Q(not_suitable='True')))
    count_prosucts_not_labeled = count_prosucts - count_prosucts_labeled
    if count_prosucts != 0:
        progress = int(count_prosucts_labeled * 100 / count_prosucts)
    else:
        progress = 0
    queryset_salad = Product.objects.filter(timetable__datetime=date_default).filter(category='Салаты')
    queryset_soup = Product.objects.filter(timetable__datetime=date_default).filter(category='Первые блюда')
    queryset_main_dishes = Product.objects.filter(timetable__datetime=date_default).filter(category='Вторые блюда')
    queryset_side_dishes = Product.objects.filter(timetable__datetime=date_default).filter(category='Гарниры')
    if request.method == 'POST' and 'save' in request.POST:
        form_date = TimetableForm(request.POST)
        formset_salad = \
            ProductFormSet(request.POST, request.FILES, queryset=queryset_salad, prefix='salad')
        formset_soup = \
            ProductFormSet(request.POST, request.FILES, queryset=queryset_soup, prefix='soup')
        formset_main_dishes = \
            ProductFormSet(request.POST, request.FILES, queryset=queryset_main_dishes, prefix='main_dishes')
        formset_side_dishes = \
            ProductFormSet(request.POST, request.FILES, queryset=queryset_side_dishes, prefix='side_dishes')

        if not formset_salad.is_valid() \
                or not formset_soup.is_valid() \
                or not formset_main_dishes.is_valid() \
                or not formset_side_dishes.is_valid():
            return render(request,
                          'index.html',
                          {'formset_salad': formset_salad,
                           'formset_soup': formset_soup,
                           'formset_main_dishes': formset_main_dishes,
                           'formset_side_dishes': formset_side_dishes,
                           'count_prosucts': count_prosucts,
                           'count_prosucts_labeled': count_prosucts_labeled,
                           'count_prosucts_not_labeled': count_prosucts_not_labeled,
                           'form_date': form_date,
                           'progress': progress,
                           })
        else:
            formset_salad.save()
            formset_soup.save()
            formset_main_dishes.save()
            formset_side_dishes.save()
            date_default = str(request.POST['datetime'])
            queryset_salad = Product.objects.filter(timetable__datetime=date_default).filter(category='Салаты')
            queryset_soup = Product.objects.filter(timetable__datetime=date_default).filter(category='Первые блюда')
            queryset_main_dishes = Product.objects.filter(timetable__datetime=date_default).filter(
                category='Вторые блюда')
            queryset_side_dishes = Product.objects.filter(timetable__datetime=date_default).filter(category='Гарниры')

    if request.method == 'POST' and 'find_date' in request.POST:
        form_date = TimetableForm(request.POST)
        if form_date.is_valid():
            date_default = str(form_date.cleaned_data["datetime"])
            queryset_salad = Product.objects.filter(timetable__datetime=str(form_date.cleaned_data["datetime"])).filter(
                category='Салаты')
            queryset_soup = Product.objects.filter(timetable__datetime=str(form_date.cleaned_data["datetime"])).filter(
                category='Первые блюда')
            queryset_main_dishes = Product.objects.filter(
                timetable__datetime=str(form_date.cleaned_data["datetime"])).filter(
                category='Вторые блюда')
            queryset_side_dishes = Product.objects.filter(
                timetable__datetime=str(form_date.cleaned_data["datetime"])).filter(category='Гарниры')

            formset_salad = ProductFormSet(queryset=queryset_salad, prefix='salad')
            formset_soup = ProductFormSet(queryset=queryset_soup, prefix='soup')
            formset_main_dishes = ProductFormSet(queryset=queryset_main_dishes, prefix='main_dishes')
            formset_side_dishes = ProductFormSet(queryset=queryset_side_dishes, prefix='side_dishes')
            data = {
                'form_date': form_date,
                'error': error,
                'formset_salad': formset_salad,
                'formset_main_dishes': formset_main_dishes,
                'formset_side_dishes': formset_side_dishes,
                'formset_soup': formset_soup,
                'count_prosucts': count_prosucts,
                'count_prosucts_labeled': count_prosucts_labeled,
                'count_prosucts_not_labeled': count_prosucts_not_labeled,
                'progress': progress,

                # 'formset_e': formset._errors,
            }
            return render(request, 'index.html', context=data)
        else:
            error = 'Некорректные данные'
    else:
        formset_salad = ProductFormSet(queryset=queryset_salad, prefix='salad')
        formset_soup = ProductFormSet(queryset=queryset_soup, prefix='soup')
        formset_main_dishes = ProductFormSet(queryset=queryset_main_dishes, prefix='main_dishes')
        formset_side_dishes = ProductFormSet(queryset=queryset_side_dishes, prefix='side_dishes')

        form_date = TimetableForm(initial={'datetime': date_default})
        data = {
            'form_date': form_date,
            'error': error,
            'formset_salad': formset_salad,
            'formset_main_dishes': formset_main_dishes,
            'formset_side_dishes': formset_side_dishes,
            'formset_soup': formset_soup,
            'count_prosucts': count_prosucts,
            'count_prosucts_labeled': count_prosucts_labeled,
            'count_prosucts_not_labeled': count_prosucts_not_labeled,
            'progress': progress,

            # 'formset_e': formset._errors
        }
    return render(request, 'index.html', context=data)


@user_passes_test(group_nutritionists_check, login_url='login')
@login_required(login_url='login')
def search(request):
    error = ''
    ProductFormSet = modelformset_factory(Product,
                                          fields=(
                                              'iditem', 'name', 'description', 'ovd', 'ovd_sugarless', 'shd', 'bd',
                                              'vbd', 'nbd', 'nkd',
                                              'vkd', 'not_suitable', 'carbohydrate', 'fat', 'fiber', 'energy', 'category',
                                              'cooking_method', 'comment'),
                                          widgets={'ovd': CheckboxInput(
                                              attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'ovd_sugarless': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'shd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'bd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'not_suitable': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'name': Textarea(attrs={'style': "display: none;"}),
                                              'description': Textarea(attrs={'style': "display: none;"}),
                                              'carbohydrate': Textarea(attrs={'style': "display: none;"}),
                                              'iditem': Textarea(attrs={'style': "display: none;"}),
                                              'fat': Textarea(attrs={'style': "display: none;"}),
                                              'fiber': Textarea(attrs={'style': "display: none;"}),
                                              'energy': Textarea(attrs={'style': "display: none;"}),
                                              'category': Textarea(attrs={'style': "display: none;"}),
                                              'cooking_method': Textarea(attrs={'style': "display: none;"}),
                                              'comment': Textarea(attrs={'class': "form-control", 'rows': "3"}),
                                          },
                                          extra=0, )
    formset = ''
    queryset = Product.objects.all()
    if request.method == 'POST':
        formset = \
            ProductFormSet(request.POST, request.FILES, queryset=queryset)
        if formset.is_valid():
            formset.save()
            queryset = Product.objects.filter(name=formset.data['form-0-name'])
            formset = ProductFormSet(queryset=queryset)
        else:
            error = 'error'
    if request.GET != {}:
        queryset = Product.objects.filter(name=request.GET['search'])
        formset = ProductFormSet(queryset=queryset)
    products = Product.objects.all()
    data = {
        'formset': formset,
        'products': products,
        'error': error,
        }
    return render(request, 'search.html', context=data)

def page_calc(page, count_prosucts):
    page = int(page)
    page_start = page - 25
    page_finish = page - 1
    page_count = math.ceil(count_prosucts / 25)
    page_dict = {item: str(item * 25) for item in range(1, page_count + 1)}
    page_prev = page - 25
    page_next = page + 25
    if page_next > (page_count * 25):
        page_next = 0
    return page_start, page_finish, str(page_prev), str(page_next), page_dict


def get_stat(category):
    count_prosucts = len(Product.objects.filter(category=category))
    count_prosucts_labeled = len(Product.objects.filter(category=category).filter(
        Q(ovd='True') | Q(ovd_sugarless='True') | Q(shd='True') | Q(bd='True') | Q(vbd='True') | Q(nbd='True') | Q(
            nkd='True') | Q(vkd='True') | Q(not_suitable='True')))
    count_prosucts_not_labeled = count_prosucts - count_prosucts_labeled
    progress = int(count_prosucts_labeled * 100 / count_prosucts)
    return count_prosucts, count_prosucts_labeled, count_prosucts_not_labeled, progress


@login_required(login_url='login')
@user_passes_test(group_nutritionists_check, login_url='login')
def catalog_salad(request, page):
    error = ''
    ProductFormSet = modelformset_factory(Product,
                                          fields=(
                                              'iditem', 'name', 'description', 'ovd', 'ovd_sugarless', 'shd', 'bd',
                                              'vbd', 'nbd', 'nkd',
                                              'vkd', 'not_suitable', 'carbohydrate', 'fat', 'fiber', 'energy', 'category',
                                              'cooking_method', 'comment'),
                                          widgets={'ovd': CheckboxInput(
                                              attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'ovd_sugarless': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'shd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'bd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'not_suitable': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'name': Textarea(attrs={'style': "display: none;"}),
                                              'description': Textarea(attrs={'style': "display: none;"}),
                                              'carbohydrate': Textarea(attrs={'style': "display: none;"}),
                                              'iditem': Textarea(attrs={'style': "display: none;"}),
                                              'fat': Textarea(attrs={'style': "display: none;"}),
                                              'fiber': Textarea(attrs={'style': "display: none;"}),
                                              'energy': Textarea(attrs={'style': "display: none;"}),
                                              'category': Textarea(attrs={'style': "display: none;"}),
                                              'cooking_method': Textarea(attrs={'style': "display: none;"}),
                                              'comment': Textarea(attrs={'class': "form-control", 'rows': "3"}),

                                          },
                                          extra=0, )
    count_prosucts, count_prosucts_labeled, count_prosucts_not_labeled, progress = get_stat('Салаты')

    page_start, page_finish, page_prev, page_next, page_dict = page_calc(page, count_prosucts)
    q = Product.objects.filter(category='Салаты').order_by(Lower('name'))[page_start:page_finish]
    queryset = Product.objects.filter(id__in=[item_q.id for item_q in q])
    queryset = queryset.order_by(Lower('name'))

    if request.method == 'POST' and 'save' in request.POST:
        formset = \
            ProductFormSet(request.POST, request.FILES, queryset=queryset, prefix='salad')
        if not formset.is_valid():
            return render(request,
                          'salad.html',
                          {'formset': formset,
                           'count_prosucts': count_prosucts,
                           'count_prosucts_labeled': count_prosucts_labeled,
                           'count_prosucts_not_labeled': count_prosucts_not_labeled,
                           'progress': progress,
                           })
        else:
            formset.save()
            count_prosucts, count_prosucts_labeled, count_prosucts_not_labeled, progress = get_stat('Салаты')
            data = {
                'page': page,
                'page_prev': page_prev,
                'page_next': page_next,
                'page_dict': page_dict,
                'formset': formset,
                'count_prosucts': count_prosucts,
                'count_prosucts_labeled': count_prosucts_labeled,
                'count_prosucts_not_labeled': count_prosucts_not_labeled,
                'progress': progress,
                # 'formset_e': formset._errors
            }
            return render(request, 'salad.html', context=data)

    else:
        formset = ProductFormSet(queryset=queryset, prefix='salad')

        data = {
            'queryset': queryset,
            'page': page,
            'page_prev': page_prev,
            'page_next': page_next,
            'page_dict': page_dict,
            'formset': formset,
            'count_prosucts': count_prosucts,
            'count_prosucts_labeled': count_prosucts_labeled,
            'count_prosucts_not_labeled': count_prosucts_not_labeled,
            'progress': progress,
            # 'formset_e': formset._errors
        }
    return render(request, 'salad.html', context=data)


@login_required(login_url='login')
@user_passes_test(group_nutritionists_check, login_url='login')
def catalog_soup(request, page):
    error = ''
    ProductFormSet = modelformset_factory(Product,
                                          fields=(
                                              'iditem', 'name', 'description', 'ovd', 'ovd_sugarless', 'shd', 'bd',
                                              'vbd', 'nbd', 'nkd',
                                              'vkd', 'not_suitable', 'carbohydrate', 'fat', 'fiber', 'energy', 'category',
                                              'cooking_method', 'comment'),
                                          widgets={'ovd': CheckboxInput(
                                              attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'ovd_sugarless': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'shd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'bd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'not_suitable': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'name': Textarea(attrs={'style': "display: none;"}),
                                              'description': Textarea(attrs={'style': "display: none;"}),
                                              'carbohydrate': Textarea(attrs={'style': "display: none;"}),
                                              'iditem': Textarea(attrs={'style': "display: none;"}),
                                              'fat': Textarea(attrs={'style': "display: none;"}),
                                              'fiber': Textarea(attrs={'style': "display: none;"}),
                                              'energy': Textarea(attrs={'style': "display: none;"}),
                                              'category': Textarea(attrs={'style': "display: none;"}),
                                              'cooking_method': Textarea(attrs={'style': "display: none;"}),
                                              'comment': Textarea(attrs={'class': "form-control", 'rows': "3"}),
                                          },
                                          extra=0, )

    count_prosucts, count_prosucts_labeled, count_prosucts_not_labeled, progress = get_stat('Первые блюда')

    page_start, page_finish, page_prev, page_next, page_dict = page_calc(page, count_prosucts)
    q = Product.objects.filter(category='Первые блюда').order_by(Lower('name'))[page_start:page_finish]
    queryset = Product.objects.filter(id__in=[item_q.id for item_q in q])
    queryset = queryset.order_by(Lower('name'))

    if request.method == 'POST' and 'save' in request.POST:
        formset = \
            ProductFormSet(request.POST, request.FILES, queryset=queryset, prefix='soup')
        if not formset.is_valid():
            return render(request,
                          'soup.html',
                          {'formset': formset,
                           'count_prosucts': count_prosucts,
                           'count_prosucts_labeled': count_prosucts_labeled,
                           'count_prosucts_not_labeled': count_prosucts_not_labeled,
                           'progress': progress,

                           })
        else:
            formset.save()
            count_prosucts, count_prosucts_labeled, count_prosucts_not_labeled, progress = get_stat('Первые блюда')
            data = {
                'page': page,
                'page_prev': page_prev,
                'page_next': page_next,
                'page_dict': page_dict,
                'formset': formset,
                'count_prosucts': count_prosucts,
                'count_prosucts_labeled': count_prosucts_labeled,
                'count_prosucts_not_labeled': count_prosucts_not_labeled,
                'progress': progress,
                # 'formset_e': formset._errors
            }
            return render(request, 'soup.html', context=data)

    else:
        formset = ProductFormSet(queryset=queryset, prefix='soup')

        data = {
            'page': page,
            'page_prev': page_prev,
            'page_next': page_next,
            'page_dict': page_dict,
            'formset': formset,
            'count_prosucts': count_prosucts,
            'count_prosucts_labeled': count_prosucts_labeled,
            'count_prosucts_not_labeled': count_prosucts_not_labeled,
            'progress': progress,
            # 'formset_e': formset._errors
        }
    return render(request, 'soup.html', context=data)


@login_required
@user_passes_test(group_nutritionists_check, login_url='login')
def catalog_main_dishes(request, page):
    # 335
    error = ''
    ProductFormSet = modelformset_factory(Product,
                                          fields=(
                                              'iditem', 'name', 'description', 'ovd', 'ovd_sugarless', 'shd', 'bd',
                                              'vbd', 'nbd', 'nkd',
                                              'vkd', 'not_suitable', 'carbohydrate', 'fat', 'fiber', 'energy', 'category',
                                              'cooking_method', 'comment'),
                                          widgets={'ovd': CheckboxInput(
                                              attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'ovd_sugarless': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'shd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'bd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'not_suitable': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'name': Textarea(attrs={'style': "display: none;"}),
                                              'description': Textarea(attrs={'style': "display: none;"}),
                                              'carbohydrate': Textarea(attrs={'style': "display: none;"}),
                                              'iditem': Textarea(attrs={'style': "display: none;"}),
                                              'fat': Textarea(attrs={'style': "display: none;"}),
                                              'fiber': Textarea(attrs={'style': "display: none;"}),
                                              'energy': Textarea(attrs={'style': "display: none;"}),
                                              'category': Textarea(attrs={'style': "display: none;"}),
                                              'cooking_method': Textarea(attrs={'style': "display: none;"}),
                                              'comment': Textarea(attrs={'class': "form-control", 'rows': "3"}),
                                          },
                                          extra=0, )
    count_prosucts, count_prosucts_labeled, count_prosucts_not_labeled, progress = get_stat('Вторые блюда')

    page_start, page_finish, page_prev, page_next, page_dict = page_calc(page, count_prosucts)
    q = Product.objects.filter(category='Вторые блюда').order_by(Lower('name'))[page_start:page_finish]
    queryset = Product.objects.filter(id__in=[item_q.id for item_q in q])
    queryset = queryset.order_by(Lower('name'))

    if request.method == 'POST' and 'save' in request.POST:
        formset = \
            ProductFormSet(request.POST, request.FILES, queryset=queryset, prefix='main_dishes')
        if not formset.is_valid():
            return render(request,
                          'main_dishes.html',
                          {'formset': formset,
                           'count_prosucts': count_prosucts,
                           'count_prosucts_labeled': count_prosucts_labeled,
                           'count_prosucts_not_labeled': count_prosucts_not_labeled,
                           'progress': progress,

                           })
        else:
            formset.save()
            count_prosucts, count_prosucts_labeled, count_prosucts_not_labeled, progress = get_stat('Вторые блюда')
            data = {
                'page': page,
                'page_prev': page_prev,
                'page_next': page_next,
                'page_dict': page_dict,
                'formset': formset,
                'count_prosucts': count_prosucts,
                'count_prosucts_labeled': count_prosucts_labeled,
                'count_prosucts_not_labeled': count_prosucts_not_labeled,
                'progress': progress,
                # 'formset_e': formset._errors
            }
            return render(request, 'main_dishes.html', context=data)

    else:
        formset = ProductFormSet(queryset=queryset, prefix='main_dishes')

        data = {
            'page': page,
            'page_prev': page_prev,
            'page_next': page_next,
            'page_dict': page_dict,
            'formset': formset,
            'count_prosucts': count_prosucts,
            'count_prosucts_labeled': count_prosucts_labeled,
            'count_prosucts_not_labeled': count_prosucts_not_labeled,
            'progress': progress,
            # 'formset_e': formset._errors
        }
    return render(request, 'main_dishes.html', context=data)


@login_required(login_url='login')
@user_passes_test(group_nutritionists_check, login_url='login')
def catalog_side_dishes(request, page):
    # 157
    error = ''
    ProductFormSet = modelformset_factory(Product,
                                          fields=(
                                              'iditem', 'name', 'description', 'ovd', 'ovd_sugarless', 'shd', 'bd',
                                              'vbd', 'nbd', 'nkd',
                                              'vkd', 'not_suitable', 'carbohydrate', 'fat', 'fiber', 'energy', 'category',
                                              'cooking_method', 'comment'),
                                          widgets={'ovd': CheckboxInput(
                                              attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'ovd_sugarless': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'shd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'bd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nbd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'nkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'vkd': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'not_suitable': CheckboxInput(
                                                  attrs={'class': 'form-check-input', 'type': 'checkbox'}),
                                              'name': Textarea(attrs={'style': "display: none;"}),
                                              'description': Textarea(attrs={'style': "display: none;"}),
                                              'carbohydrate': Textarea(attrs={'style': "display: none;"}),
                                              'iditem': Textarea(attrs={'style': "display: none;"}),
                                              'fat': Textarea(attrs={'style': "display: none;"}),
                                              'fiber': Textarea(attrs={'style': "display: none;"}),
                                              'energy': Textarea(attrs={'style': "display: none;"}),
                                              'category': Textarea(attrs={'style': "display: none;"}),
                                              'cooking_method': Textarea(attrs={'style': "display: none;"}),
                                              'comment': Textarea(attrs={'class': "form-control", 'rows': "3"}),
                                          },
                                          extra=0, )
    count_prosucts, count_prosucts_labeled, count_prosucts_not_labeled, progress = get_stat('Гарниры')

    page_start, page_finish, page_prev, page_next, page_dict = page_calc(page, count_prosucts)
    q = Product.objects.filter(category='Гарниры').order_by(Lower('name'))[page_start:page_finish]
    queryset = Product.objects.filter(id__in=[item_q.id for item_q in q])
    queryset = queryset.order_by(Lower('name'))

    if request.method == 'POST' and 'save' in request.POST:
        formset = \
            ProductFormSet(request.POST, request.FILES, queryset=queryset, prefix='side_dishes')
        if not formset.is_valid():
            return render(request,
                          'side_dishes.html',
                          {'formset': formset,
                           'count_prosucts': count_prosucts,
                           'count_prosucts_labeled': count_prosucts_labeled,
                           'count_prosucts_not_labeled': count_prosucts_not_labeled,
                           'progress': progress,

                           })
        else:
            formset.save()
            count_prosucts, count_prosucts_labeled, count_prosucts_not_labeled, progress = get_stat('Гарниры')
            data = {
                'page': page,
                'page_prev': page_prev,
                'page_next': page_next,
                'page_dict': page_dict,
                'formset': formset,
                'formset': formset,
                'count_prosucts': count_prosucts,
                'count_prosucts_labeled': count_prosucts_labeled,
                'count_prosucts_not_labeled': count_prosucts_not_labeled,
                'progress': progress,
                # 'formset_e': formset._errors
            }
            return render(request, 'side_dishes.html', context=data)

    else:
        formset = ProductFormSet(queryset=queryset, prefix='side_dishes')

        data = {
            'page': page,
            'page_prev': page_prev,
            'page_next': page_next,
            'page_dict': page_dict,
            'formset': formset,
            'formset': formset,
            'count_prosucts': count_prosucts,
            'count_prosucts_labeled': count_prosucts_labeled,
            'count_prosucts_not_labeled': count_prosucts_not_labeled,
            'progress': progress,
            # 'formset_e': formset._errors
        }
    return render(request, 'side_dishes.html', context=data)


# class BaseAPIView(APIView):
#     def post(self, request):
#         data = request.data
#         data_str = str(data)
#         data_dict = dict(data)
#         js = open("cooking_method.json").read()
#         dict_tk = json.loads(js)
#         if data_dict['menu']['location']['name'] == 'hadassah':
#             load_menu(data_dict, dict_tk)
#             load_timetable(data_dict)
#             Base.objects.create(base=data_str)
#         return Response(data)

class BaseAPIView(APIView):
    def post(self, request):
        data = request.data
        data_str = str(data)
        Base.objects.create(base=data_str)
        data_dict = dict(data)
        js = open("cooking_method.json").read()
        dict_tk = json.loads(js)
        if data_dict['menu']['location']['name'] == 'hadassah':
            load_menu(data_dict, dict_tk)
            load_timetable(data_dict)
            Base.objects.create(base=data_str)
        return Response(data)


class VerifyAPIView(APIView):
    def post(self, request):
        data = request.data
        try:
            if Barcodes.objects.get(number=data['barcode']):
                return Response('No')
        except Exception:
            pass
        return Response('Yes')



class DeactivateAPIView(APIView):
    def post(self, request):
        data = request.data
        try:
            Barcodes.objects.get(number=data['barcode'])
            return Response('Barcode already deactivated')
        except Exception:
            Barcodes(number=data['barcode'], status='no_active').save()
            return Response('Yes')


def user_login(request):
    errors = None
    if request.method == 'POST':
        user_form = UserloginForm(request.POST)
        user = authenticate(username=user_form.data['username'],
                            password=user_form.data['password'])
        if user is not None:
            login(request, user)
            if user.groups.filter(name='nutritionists').exists():
                return HttpResponseRedirect(reverse('index'))
            if user.groups.filter(name='doctors').exists():
                return HttpResponseRedirect(reverse('doctor'))
            if user.groups.filter(name='patients').exists():
                return HttpResponseRedirect(reverse('patient'))
        else:
            errors = 'Пользователь с таким именем или паролем не существует'
    else:
        user_form = UserloginForm()
    return render(request, 'nutritionist/registration/login.html', {'user_form': user_form,
                                                       'errors': errors})


def user_logout(request):
    from django.contrib.auth import logout
    logout(request)
    return HttpResponseRedirect(reverse('login'))


def register(request):
    """ Регистрация нового пользователя"""
    errors = []
    if request.method == 'POST':
        password = ''.join([random.choice("123456789qwertyuiopasdfghjklzxcvbnm") for i in range(5)])
        user_form = UserRegistrationForm(request.POST)
        if user_form.is_valid():
            try:
                user = CustomUser.objects.create_user(user_form.data['email'], user_form.data['email'], password)
                # def chk_rest(s):
                #     pat = r'([^а-яА-Яa-zA-Z0-9ёЁ\)\(-\' .,])'
                #     return re.findall(pat, flags=re.I)
                user.first_name = user_form.data['name']
                user.last_name = user_form.data['lastname']
                group_nutritionists = Group.objects.get(name='doctors')
                group_nutritionists.user_set.add(user)
                user.save()
                text_email = f"<p>Регистрация прошла успешно!</p>\
                            <p>логин: {user_form.data['email']}</p> \
                            <p>пароль: {password}</p> \
                            <p>Для авторизации пройдите по ссылке \
                            <a href='https://sk.petrushkagroup.com/login/'>sk.petrushkagroup.com/login/</a></p>"

                send_mail(
                    'Регистрация в личном кабинете врача.',
                    text_email,
                    'info@petrushkagroup.com',
                    [user_form.data['email']],
                    fail_silently=False,
                    html_message=text_email,
                )

                return render(request, 'nutritionist/registration/register_done.html', {'user_form': user_form, 'errors': errors})
            except Exception:
                errors.append('Пользователь с такой почтой уже существует')
                return render(request, 'nutritionist/registration/register.html',
                              {'user_form': user_form, 'errors': errors})


    user_form = UserRegistrationForm()
    return render(request, 'nutritionist/registration/register.html', {'user_form': user_form, 'errors': errors})


def password_reset(request):
    """ Сброс пароля """
    errors = []
    password = ''.join([random.choice("123456789qwertyuiopasdfghjklzxcvbnm") for i in range(5)])
    if request.method == 'POST':
        user_form = UserPasswordResetForm(request.POST)
        if user_form.is_valid():
            try:
                user = CustomUser.objects.get(email=user_form.data['email'])
                user.set_password(password)
                user.save()
                text_email = f"<p>Сброс пароля прошел успешно!</p>\
                            <p>логин: {user_form.data['email']}</p> \
                            <p>пароль: {password}</p> \
                            <p>Для авторизации пройдите по ссылке \
                            <a href='https://sk.petrushkagroup.com/login/'>sk.petrushkagroup.com/login/</a></p>"
                send_mail(
                    'Сброс пароля',
                    text_email,
                    'info@petrushkagroup.com',
                    [user_form.data['email']],
                    fail_silently=False,
                    html_message=text_email,
                )
                return render(request, 'nutritionist/registration/password_reset_done.html', {'user_form': user_form, 'errors': errors})
            except Exception:
                errors.append('Пользователь с такой почтой не зарегистророван')
    else:
        user_form = UserPasswordResetForm()
    return render(request, 'nutritionist/registration/password_reset_email.html', {'user_form': user_form, 'errors': errors})

def manager(request):
    data = {}
    return render(request, 'admin.html', context=data)

def printed_form_one(request):

    formatted_date_now = dateformat.format(date.fromisoformat(str(date.today())), 'd E, l')
    floors = {
    'second': ['2а-1', '2а-2', '2а-3', '2а-4', '2а-5', '2а-6', '2а-7', '2а-12', '2а-13', '2а-14', '2а-15',
                     '2а-16', '2а-17'],
    'third': ['3а-1', '3а-2', '3а-3', '3а-4', '3а-5', '3а-6', '3а-7', '3а-8', '3а-9', '3а-10', '3а-11',
                    '3а-12', '3а-13', '3а-14', '3а-15', '3а-16', '3а-17'],
    'fourtha': ['4а-1', '4а-2', '4а-3', '4а-4', '4а-5', '4а-6', '4а-7', '4а-8', '4а-9', '4а-10', '4а-11',
                      '4а-12', '4а-13', '4а-14', '4а-15', '4а-16'],
    }
    time_now = str(datetime.today().time().hour) + ':' + str(datetime.today().time().minute)
    # какой прием пищи
    meal, day = what_meal() # после return 'breakfast', 'tomorrow'
    type_order = what_type_order()
    date_create = date.today() + timedelta(days=1) if day == 'tomorrow' else date.today()
    if type_order == 'flex-order':
        users = UsersToday.objects.all()
    else:
        users = UsersReadyOrder.objects.all()

    catalog = {'meal': translate_meal(meal),
               'count': len(users),
               'count_2nd_floor': len([user for user in users if user.room_number in floors['second']]),
               'count_3nd_floor': len([user for user in users if user.room_number in floors['third']]),
               'count_4nd_floor': len([user for user in users if user.room_number in floors['fourtha']]),
               'count_diet': counting_diets(users, floors),
               'users_2nd_floor': create_list_users_on_floor(users, floors['second'], meal, date_create, type_order),
               'users_3nd_floor': create_list_users_on_floor(users, floors['third'], meal, date_create, type_order),
               'users_4nd_floor': create_list_users_on_floor(users, floors['fourtha'], meal, date_create, type_order),
               }
    number = 0
    count_users_with_cafe_prod = 0
    # расставляем порядковые номера и считаем кол-во пациетнов с блюдами с кафе
    for user in catalog['users_2nd_floor']:
        number += 1
        user['number'] = str(number)
        if len(user['products_cafe']) > 0:
            count_users_with_cafe_prod += 1
    for user in catalog['users_3nd_floor']:
        number += 1
        user['number'] = str(number)
        if len(user['products_cafe']) > 0:
            count_users_with_cafe_prod += 1
    for user in catalog['users_4nd_floor']:
        number += 1
        user['number'] = str(number)
        if len(user['products_cafe']) > 0:
            count_users_with_cafe_prod += 1

    data = {
        'formatted_date': formatted_date_now,
        'time_now': time_now,
        'catalog': catalog,
        'count_users_with_cafe_prod': count_users_with_cafe_prod,
        'day': day,
        'date_create': dateformat.format(date.fromisoformat(str(date_create)), 'd E')
    }

    return render(request, 'printed_form1.html', context=data)


def printed_form_two_lp(request):
    formatted_date_now = dateformat.format(date.fromisoformat(str(date.today())), 'd E, l')
    time_now = str(datetime.today().time().hour) + ':' + str(datetime.today().time().minute)
    # какой прием пищи
    meal, day = what_meal() # после return 'breakfast', 'tomorrow'
    type_order = what_type_order()
    date_create = date.today() + timedelta(days=1) if day == 'tomorrow' else date.today()
    catalog = {}

    if type_order == 'flex-order':
        users = UsersToday.objects.all()
    else:
        users = UsersReadyOrder.objects.all()
    # users = users.filter(status='patient').filter(receipt_date__lte=date.today())
    for category in ['porridge', 'salad', 'soup', 'main', 'garnish', 'dessert', 'fruit', 'drink', 'products']:
        list_whith_unique_products = []
        for diet in ['ОВД', 'ОВД без сахара', 'ЩД', 'ОВД веган (пост) без глютена', 'Нулевая диета', 'БД', 'ВБД', 'НБД', 'НКД', 'ВКД', 'БД день 1', 'БД день 2']:
            users_with_diet = users.filter(type_of_diet=diet)
            all_products = []
            for user in users_with_diet:
                if type_order == 'flex-order':
                    menu_all = MenuByDay.objects.filter(user_id=user.user_id)
                else:
                    menu_all = MenuByDayReadyOrder.objects.filter(user_id=user.id)
                if category == 'products' or category ==  'drink':
                    pr = check_value_two(menu_all, str((date_create)), meal, category)
                else:
                    pr = [check_value_two(menu_all, str((date_create)), meal, category)]
                # if pr != None:
                #     all_products.append(pr)
                if pr[0] != None:
                    for item in pr:
                        all_products.append(item)
            # составляем список с уникальными продуктами
            unique_products = []
            for product in all_products:
                flag = True
                for un_product in unique_products:
                    if product != None or un_product != None:
                        if product['id'] == un_product['id']:
                            flag = False
                if flag == True:
                    if product:
                        if 'cafe' not in product['id']:
                            unique_products.append(product)
            # добавляем элеметны списока с уникальными продуктами, кол-вом(сколько продуктов всего)
            # типом диеты
            for un_product in unique_products:
                count = 0
                for product in all_products:
                    if product['id'] == un_product['id']:
                        count += 1
                un_product['count'] = str(count)
                un_product['diet'] = diet
            # list_whith_unique_products.append(unique_products)
            [list_whith_unique_products.append(item) for item in unique_products]
        catalog[category] = list_whith_unique_products

    for cat in catalog.values():
        for i, pr in enumerate(cat):
            for ii in range(i + 1, len(cat)):
                if pr != None and cat[ii] != None:
                    if pr['id'] == cat[ii]['id']:
                        pr['count'] = str(int(pr['count']) + int(cat[ii]['count']))
                        pr['diet'] = pr['diet'] + ', ' + cat[ii]['diet']
                        cat[ii] = None

    number = 1
    for item in catalog.values():
        for product in item:
            if product != None:
                product['number'] = number
                number += 1



    data = {
        'formatted_date': formatted_date_now,
        'time_now': time_now,
        'catalog': catalog,
        'day': day,
        'date_create': date_create,
        'meal': translate_meal(meal)
    }
    return render(request, 'printed_form2_lp.html', context=data)


def printed_form_two_cafe(request):
    formatted_date_now = dateformat.format(date.fromisoformat(str(date.today())), 'd E, l')
    time_now = str(datetime.today().time().hour) + ':' + str(datetime.today().time().minute)
    # какой прием пищи
    meal, day = what_meal() # после return 'breakfast', 'tomorrow'
    type_order = what_type_order()
    # type_order = 'fix-order'
    date_create = date.today() + timedelta(days=1) if day == 'tomorrow' else date.today()
    # meal = 'lunch'
    catalog = {}

    # users = CustomUser.objects.all()
    # users = users.filter(status='patient').filter(receipt_date__lte=date.today())
    if type_order == 'flex-order':
        users = UsersToday.objects.all()
    else:
        users = UsersReadyOrder.objects.all()
    for category in ['porridge', 'salad', 'soup', 'main', 'garnish', 'dessert', 'fruit', 'drink']:
        list_whith_unique_products = []
        for diet in ['ОВД', 'ОВД без сахара', 'ОВД веган (пост) без глютена', 'Нулевая диета', 'ЩД', 'БД', 'БД день 1', 'БД день 2', 'ВБД', 'НБД', 'НКД', 'ВКД']:
            users_with_diet = users.filter(type_of_diet=diet)
            all_products = []
            for user in users_with_diet:
                if type_order == 'flex-order':
                    menu_all = MenuByDay.objects.filter(user_id=user.user_id)
                else:
                    menu_all = MenuByDayReadyOrder.objects.filter(user_id=user.id)
                all_products.append(check_value_two(menu_all, str(date.today()), meal, category))
            # составляем список с уникальными продуктами
            unique_products = []
            for product in all_products:
                flag = True
                for un_product in unique_products:
                    if product != None or un_product != None:
                        if product['id'] == un_product['id']:
                            flag = False
                if flag == True:
                    if product:
                        if 'cafe' in product['id']:
                            unique_products.append(product)
            # добавляем элеметны списока с уникальными продуктами, кол-вом(сколько продуктов всего)
            # типом диеты
            for un_product in unique_products:
                count = 0
                for product in all_products:
                    if product['id'] == un_product['id']:
                        count += 1
                un_product['count'] = str(count)
                un_product['diet'] = diet
            # list_whith_unique_products.append(unique_products)
            [list_whith_unique_products.append(item) for item in unique_products]
        catalog[category] = list_whith_unique_products

    for cat in catalog.values():
        for i, pr in enumerate(cat):
            for ii in range(i + 1, len(cat)):
                if pr != None and cat[ii] != None:
                    if pr['id'] == cat[ii]['id']:
                        pr['count'] = str(int(pr['count']) + int(cat[ii]['count']))
                        pr['diet'] = pr['diet'] + ', ' + cat[ii]['diet']
                        cat[ii] = None

    number = 1
    for item in catalog.values():
        for product in item:
            if product != None:
                product['number'] = number
                number += 1

    data = {
        'formatted_date': formatted_date_now,
        'time_now': time_now,
        'catalog': catalog,
        'day': day,
        'date_create': date_create,
        'meal': translate_meal(meal)
    }
    return render(request, 'printed_form2_cafe.html', context=data)


def create_external_report(filtered_report):
    report = {}
    report_ = {}
    for index, item in enumerate(filtered_report):
        report.setdefault(str(item.date_create), []).append(item)

    for key, value in report.items():
        report_[key] = {}
        for index, item in enumerate(value):
            report_[key].setdefault(str(item.meal), []).append(item)

    for key1, value1 in report_.items():
        report[key1] = {}
        count = 0
        for key2, value2 in value1.items():
            report[key1][key2] = {}
            for index, item in enumerate(value2):
                report[key1][key2].setdefault(str(item.type_of_diet), []).append(item)
                test_dict = {}
            for key3, value3 in report[key1][key2].items():
                report[key1][key2][key3] = len(set([user.user_id for user in (report[key1][key2][key3])]))
                count += report[key1][key2][key3]
        report[key1]['Всего'] = {'count': count}


    return report


def get_report(report, report_detailing, date_start, date_finish):
    """ Создаёт excel фаил с отчетом по блюдам """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    font = Font(name='Arial',
                size=10,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')
    font_white = Font(name='Arial',
                size=10,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='ffffff')
    font_10_bold = Font(name='Arial',
                size=10,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')
    ws['A1'].font = font
    ws['A2'].font = font
    ws['A3'].font = font
    ws['A5'].font = font
    ws['B5'].font = font
    ws['C5'].font = font
    ws['D5'].font = font
    ws['E5'].font = font
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    ws.merge_cells('A3:E3')
    ws.merge_cells('A4:E4')
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.row_dimensions[5].height = 30
    ws['A1'].value = 'Отчет по лечебному питанию'
    ws['A1'].font = Font(name='Arial',
                         size=14,
                         bold=False,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')
    ws['A2'].value = 'Круглосуточный стационар, Hadassah Medical Moscow'
    ws['A2'].font = Font(name='Arial',
                         size=11,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')
    ws['A3'].value = f'{date_start.day}.{date_start.month}.{date_start.year} - {date_finish.day}.{date_finish.month}.{date_finish.year}'
    ws['A3'].font = Font(name='Arial',
                         size=11,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')

    ws['A5'].value = 'Учетный день'
    ws['B5'].value = 'Прием пищи'
    ws['C5'].value = 'Рацион'
    ws['D5'].value = 'Количество'
    ws['E5'].value = 'Сумма, руб.'
    ws['A5'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B5'].alignment = Alignment(horizontal="left", vertical="center")
    ws['C5'].alignment = Alignment(horizontal="left", vertical="center")
    ws['D5'].alignment = Alignment(horizontal="left", vertical="center")
    ws['E5'].alignment = Alignment(horizontal="left", vertical="center")

    dotted = Side(border_style="dotted", color="383636")
    thick = Side(border_style="thin", color="383636")
    row = 5
    for key1, item1 in report.items():
        row += 1
        _ = ws.cell(column=1, row=row, value=key1).font = font
        for key2, item2 in item1.items():
            if key2 == 'Всего':
                _ = ws.cell(column=2, row=row, value=key2).font = font_10_bold
            if key2 == 'breakfast':
                _ = ws.cell(column=2, row=row, value='Завтрак').font = font
            if key2 == 'lunch':
                _ = ws.cell(column=2, row=row, value='Обед').font = font
            if key2 == 'afternoon':
                _ = ws.cell(column=2, row=row, value='Полдник').font = font
            if key2 == 'dinner':
                _ = ws.cell(column=2, row=row, value='Ужин').font = font
            for key3, item3 in item2.items():
                if key2 == 'Всего':
                    _ = ws.cell(column=3, row=row, value='').font = font_10_bold
                    _ = ws.cell(column=4, row=row, value=str(item3)).font = font_10_bold
                    _ = ws.cell(column=5, row=row, value=f'{item3 * 750}.00').font = font_10_bold
                else:
                    _ = ws.cell(column=3, row=row, value=key3).font = font
                    _ = ws.cell(column=4, row=row, value=str(item3)).font = font
                    _ = ws.cell(column=5, row=row, value=f'{item3 * 750}.00').font = font
                    row += 1
            ws[row-1][0].border = Border(bottom=dotted)
            ws[row-1][1].border = Border(bottom=dotted)
            ws[row-1][2].border = Border(bottom=dotted)
            ws[row-1][3].border = Border(bottom=dotted)
            ws[row-1][4].border = Border(bottom=dotted)
        ws[row][0].border = Border(bottom=thick)
        ws[row][1].border = Border(bottom=thick)
        ws[row][2].border = Border(bottom=thick)
        ws[row][3].border = Border(bottom=thick)
        ws[row][4].border = Border(bottom=thick)

    for row in range(1, row+300):
        for col in range(1, 50):
            ws.cell(column=col, row=row).fill = PatternFill('solid', fgColor="ffffff")

    ws['A5'].fill = PatternFill('solid', fgColor="203864")
    ws['B5'].fill = PatternFill('solid', fgColor="203864")
    ws['C5'].fill = PatternFill('solid', fgColor="203864")
    ws['D5'].fill = PatternFill('solid', fgColor="203864")
    ws['E5'].fill = PatternFill('solid', fgColor="203864")
    ws['A5'].font = font_white
    ws['B5'].font = font_white
    ws['C5'].font = font_white
    ws['D5'].font = font_white
    ws['E5'].font = font_white

    ws1 = wb.create_sheet("Детальный")

    font = Font(name='Arial',
                size=10,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')
    ws1['A1'].font = font
    ws1['A2'].font = font
    ws1['A3'].font = font
    ws1['A5'].font = font
    ws1['B5'].font = font
    ws1['C5'].font = font
    ws1['D5'].font = font
    ws1['E5'].font = font
    ws1.merge_cells('A1:E1')
    ws1.merge_cells('A2:E2')
    ws1.merge_cells('A3:E3')
    ws1.merge_cells('A4:E4')
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 15
    ws1.row_dimensions[5].height = 30
    ws1['A1'].value = 'Детализация к отчету по лечебному питанию'
    ws1['A1'].font = Font(name='Arial',
                         size=14,
                         bold=False,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')
    ws1['A2'].value = 'Круглосуточный стационар, Hadassah Medical Moscow'
    ws1['A2'].font = Font(name='Arial',
                         size=11,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')
    ws1['A3'].value = f'{date_start.day}.{date_start.month}.{date_start.year} - {date_finish.day}.{date_finish.month}.{date_finish.year}'
    ws1['A3'].font = Font(name='Arial',
                         size=11,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')

    ws1['A5'].value = 'Учетный день'
    ws1['B5'].value = 'Прием пищи'
    ws1['C5'].value = 'Рацион'
    ws1['D5'].value = 'ФИО'
    ws1['E5'].value = '№ палаты'
    ws1['A5'].alignment = Alignment(horizontal="left", vertical="center")
    ws1['B5'].alignment = Alignment(horizontal="left", vertical="center")
    ws1['C5'].alignment = Alignment(horizontal="left", vertical="center")
    ws1['D5'].alignment = Alignment(horizontal="left", vertical="center")
    ws1['E5'].alignment = Alignment(horizontal="left", vertical="center")
    row = 6
    for key1, item1 in report_detailing.items():
        for key2, item2 in item1.items():
            for key3, item3 in item2.items():
                for key4, value4 in item3.items():
                    _ = ws1.cell(column=1, row=row, value=key1).font = font
                    _ = ws1.cell(column=1, row=row, value=key1).font = font
                    if key2 == 'breakfast':
                        _ = ws1.cell(column=2, row=row, value='Завтрак').font = font
                    if key2 == 'lunch':
                        _ = ws1.cell(column=2, row=row, value='Обед').font = font
                    if key2 == 'afternoon':
                        _ = ws1.cell(column=2, row=row, value='Полдник').font = font
                    if key2 == 'dinner':
                        _ = ws1.cell(column=2, row=row, value='Ужин').font = font
                    _ = ws1.cell(column=3, row=row, value=key3).font = font
                    _ = ws1.cell(column=4, row=row, value=key4).font = font
                    _ = ws1.cell(column=5, row=row, value=value4).font = font
                    row += 1


    for row in range(1, row+300):
        for col in range(1, 50):
            ws1.cell(column=col, row=row).fill = PatternFill('solid', fgColor="ffffff")


    ws1['A5'].fill = PatternFill('solid', fgColor="203864")
    ws1['B5'].fill = PatternFill('solid', fgColor="203864")
    ws1['C5'].fill = PatternFill('solid', fgColor="203864")
    ws1['D5'].fill = PatternFill('solid', fgColor="203864")
    ws1['E5'].fill = PatternFill('solid', fgColor="203864")
    ws1['A5'].font = font_white
    ws1['B5'].font = font_white
    ws1['C5'].font = font_white
    ws1['D5'].font = font_white
    ws1['E5'].font = font_white
    wb.save("static/report.xlsx")
    return

def create_external_report_detailing(filtered_report):
    report = {}
    report_ = {}
    for index, item in enumerate(filtered_report):
        report.setdefault(str(item.date_create), []).append(item)

    for key, value in report.items():
        report_[key] = {}
        for index, item in enumerate(value):
            report_[key].setdefault(str(item.meal), []).append(item)

    for key1, value1 in report_.items():
        report[key1] = {}

        for key2, value2 in value1.items():
            report[key1][key2] = {}
            for index, item in enumerate(value2):
                report[key1][key2].setdefault(str(item.type_of_diet), []).append(item)
            for key3, value3 in report[key1][key2].items():
                test = {}
                for item in report[key1][key2][key3]:
                    test[item.user_id.full_name] = item.user_id.room_number
                report[key1][key2][key3] = test
    return report


def report(request):
        # if request.method == 'GET' and request.GET != {}:
        #     date_start = parse(request.GET['start'])
        #     date_finish = parse(request.GET['finish'])
        # else:
        #     date_start = datetime(datetime.today().year, datetime.today().month, 1).date()
        #     date_finish = datetime.today().date()
        # filtered_report = Report.objects.filter(date_create__gte=date_start, date_create__lte=date_finish)
        #
        # report = {}
        # for index, item in enumerate(filtered_report):
        #     if 'cafe' in item.product_id:
        #         product = Product.objects.get(id=item.product_id.split('-')[2])
        #     else:
        #         product = ProductLp.objects.get(id=item.product_id)
        #     report.setdefault(item.product_id, []).append(
        #         {'category': product.category,
        #         'name': product.name,
        #     })
        #
        # temporary_report = []
        # for item in report.values():
        #     item[0]['count'] = len(item)
        #     temporary_report.append(item[0])
        #
        #
        # temporary_report.sort(key=operator.itemgetter('category'))
        # category = ['гарнир', 'десерт', 'напиток', 'основной', 'салат', 'суп', 'фрукты', 'каша']
        # intermediate_option = []
        # report = []
        # for cat in category:
        #     for item in temporary_report:
        #         if item['category'] == cat:
        #             intermediate_option.append(item)
        #     intermediate_option.sort(key=operator.itemgetter('name'))
        #     report += intermediate_option
        #     intermediate_option = []
        # for index, item in enumerate(report):
        #     item['category'] = item['category'] if item['category'] != 'основной' else 'основное'
        #     item['number'] = index + 1
        #
        # date = {'report': report,
        #         'date_start': date_start,
        #         'date_finish': date_finish}
    return render(request, 'report.html', {})


def reports(request):
    if request.method == 'GET' and request.GET != {}:
        date_start = parse(request.GET['start'])
        date_finish = parse(request.GET['finish'])
    else:
        date_start = datetime(datetime.today().year, datetime.today().month, 1).date()
        date_finish = datetime.today().date()
    filtered_report = Report.objects.filter(date_create__gte=date_start, date_create__lte=date_finish)

    report = {}
    for index, item in enumerate(filtered_report):
        if 'cafe' in item.product_id:
            product = Product.objects.get(id=item.product_id.split('-')[2])
        else:
            product = ProductLp.objects.get(id=item.product_id)
        report.setdefault(item.product_id, []).append(
            {'category': product.category,
            'name': product.name,
        })

    temporary_report = []
    for item in report.values():
        item[0]['count'] = len(item)
        temporary_report.append(item[0])


    temporary_report.sort(key=operator.itemgetter('category'))
    category = ['гарнир', 'десерт', 'напиток', 'основной', 'салат', 'суп', 'фрукты', 'каша']
    intermediate_option = []
    report = []
    for cat in category:
        for item in temporary_report:
            if item['category'] == cat:
                intermediate_option.append(item)
        intermediate_option.sort(key=operator.itemgetter('name'))
        report += intermediate_option
        intermediate_option = []
    for index, item in enumerate(report):
        item['category'] = item['category'] if item['category'] != 'основной' else 'основное'
        item['number'] = index + 1

    date = {'report': report,
            'date_start': date_start,
            'date_finish': date_finish}
    return render(request, 'reports.html', context=date)



class DownloadReportAPIView(APIView):
    def post(self, request):
        data = request.data
        date_start = parse(data['start'])
        date_finish = parse(data['finish'])

        filtered_report = Report.objects.filter(date_create__gte=date_start, date_create__lte=date_finish)
        report = create_external_report(filtered_report)
        report_detailing = create_external_report_detailing(filtered_report)
        get_report(report, report_detailing, date_start, date_finish)
        response = {"response": "yes"}
        response = json.dumps(response)
        return Response(response)

def create_сatalog():
    """ Создание словаря этикеток. """
    formatted_date_now = dateformat.format(date.fromisoformat(str(date.today())), 'd E, l')
    floors = {
    'second': ['2а-1', '2а-2', '2а-3', '2а-4', '2а-5', '2а-6', '2а-7', '2а-12', '2а-13', '2а-14', '2а-15',
                     '2а-16', '2а-17'],
    'third': ['3а-1', '3а-2', '3а-3', '3а-4', '3а-5', '3а-6', '3а-7', '3а-8', '3а-9', '3а-10', '3а-11',
                    '3а-12', '3а-13', '3а-14', '3а-15', '3а-16', '3а-17'],
    'fourtha': ['4а-1', '4а-2', '4а-3', '4а-4', '4а-5', '4а-6', '4а-7', '4а-8', '4а-9', '4а-10', '4а-11',
                      '4а-12', '4а-13', '4а-14', '4а-15', '4а-16'],
    }
    time_now = str(datetime.today().time().hour) + ':' + str(datetime.today().time().minute)
    # какой прием пищи
    meal, day = what_meal() # после return 'breakfast', 'tomorrow'
    type_order = what_type_order()
    date_create = date.today() + timedelta(days=1) if day == 'tomorrow' else date.today()
    if type_order == 'flex-order':
        users = UsersToday.objects.all()
    else:
        users = UsersReadyOrder.objects.all()

    catalog = {'meal': translate_meal(meal),
               'count': len(users),
               'count_2nd_floor': len([user for user in users if user.room_number in floors['second']]),
               'count_3nd_floor': len([user for user in users if user.room_number in floors['third']]),
               'count_4nd_floor': len([user for user in users if user.room_number in floors['fourtha']]),
               'count_diet': counting_diets(users, floors),
               'users_2nd_floor': create_list_users_on_floor(users, floors['second'], meal, date_create, type_order),
               'users_3nd_floor': create_list_users_on_floor(users, floors['third'], meal, date_create, type_order),
               'users_4nd_floor': create_list_users_on_floor(users, floors['fourtha'], meal, date_create, type_order),
               }
    number = 0
    count_users_with_cafe_prod = 0
    # расставляем порядковые номера и считаем кол-во пациетнов с блюдами с кафе
    for user in catalog['users_2nd_floor']:
        number += 1
        user['number'] = str(number)
        if len(user['products_cafe']) > 0:
            count_users_with_cafe_prod += 1
    for user in catalog['users_3nd_floor']:
        number += 1
        user['number'] = str(number)
        if len(user['products_cafe']) > 0:
            count_users_with_cafe_prod += 1
    for user in catalog['users_4nd_floor']:
        number += 1
        user['number'] = str(number)
        if len(user['products_cafe']) > 0:
            count_users_with_cafe_prod += 1

    data = {
        'formatted_date': formatted_date_now,
        'time_now': time_now,
        'catalog': catalog,
        'count_users_with_cafe_prod': count_users_with_cafe_prod,
        'day': day,
        'date_create': dateformat.format(date.fromisoformat(str(date_create)), 'd E')
    }
    return data


def create_stickers_pdf(catalog):
    from fpdf import FPDF

    def create_res_list(product, max_count_in_line, type):
        st_ch1 = '- ' if type == 'products' else '  '
        st_ch2 = '-' if type == 'products' else ' '
        product_list = product.split(' ')
        res = ""
        res_list = []
        for item in product_list:
            if len(res + item) < max_count_in_line:
                res = res + ' ' + item if res != "" else st_ch1 + item
            else:
                res_list.append(res) if res[0] == st_ch2 else res_list.append('  ' + res)
                res = item
        res_list.append(res) if res[0] == st_ch2 else res_list.append('  ' + res)
        return res_list

    pdf = FPDF()
    catalog = catalog['catalog']
    for floor in ['users_2nd_floor', 'users_3nd_floor', 'users_4nd_floor']:
        for item in catalog[floor]:
            pdf.set_left_margin(0)
            pdf.set_right_margin(0)
            pdf.add_page()
            # pdf.add_font("Arial", "", "FontsFree-Net-arial-bold.ttf", uni=True)
            pdf.add_font("Arial1", "", "FontsFree-Net-arial-bold.ttf", uni=True)
            pdf.set_font("Arial1", style='', size=33)
            ln = 1
            pdf.cell(50, 14, txt=f'  {formatting_full_name(item["name"])}, {item["room_number"]}, {item["bed"]}', ln=ln, align="L")
            ln += 1
            if item["diet"] == "ОВД веган (пост) без глютена":
                pdf.cell(50, 14, txt=f'  ОВД веган (пост) без глютена,', ln=ln, align="L")
                ln += 1
                pdf.cell(50, 14, txt=f'  {catalog["meal"].lower()}, {date.today().day}/{date.today().month}/{str(date.today().year)[2:]}', ln=ln, align="L")
                ln += 1
            else:
                pdf.cell(50, 14, txt=f'  {item["diet"]}, {catalog["meal"].lower()}, {date.today().day}/{date.today().month}/{str(date.today().year)[2:]}', ln=3, align="L")
                ln += 1
            pdf.add_font("Arial2", "", "arial.ttf", uni=True)
            pdf.set_font("Arial2", style='', size=30)
            max_count_in_line = 38
            if len(item["comment"]) >= max_count_in_line:
                res_list = create_res_list(item["comment"], max_count_in_line, 'comment')
                for product in res_list:
                    pdf.cell(50, 10, txt=f'{product}', ln=ln, align="L")
                    ln += 1
            else:
                pdf.cell(50, 10, txt=f'  {item["comment"]}', ln=ln, align="L")
            ln += 1
            pdf.cell(50, 10, txt="", ln=5, align="L")
            ln += 1



            for index, product in enumerate(item['products_lp'] + item['products_cafe']):
                if len(product) >= max_count_in_line:
                    res_list = create_res_list(product, max_count_in_line, 'products')
                    for product in res_list:
                        pdf.cell(50, 10, txt=f'{product}', ln=index + ln, align="L")
                        ln += 1
                else:
                    pdf.cell(50, 10, txt=f'- {product}', ln=index + ln, align="L")
                pdf.cell(50, 3, txt=f'', ln=index + 1 + ln, align="L")
    pdf.output("static/stickers.pdf")
    return


class CreateSitckers(APIView):
    def post(self, request):
        data = create_сatalog()
        create_stickers_pdf(data)
        response = {"response": "yes"}
        response = json.dumps(response)
        return Response(response)
